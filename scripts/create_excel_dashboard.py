"""Generates a professional Excel dashboard using openpyxl."""

import pathlib
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def main() -> None:
    base_dir = pathlib.Path(__file__).resolve().parent.parent
    processed_orders_path = base_dir / "data" / "processed" / "processed_orders.csv"
    excel_dir = base_dir / "excel"
    excel_dir.mkdir(parents=True, exist_ok=True)
    excel_path = excel_dir / "Promotion_Analytics_Dashboard.xlsx"
    
    if not processed_orders_path.exists():
        print(f"Processed orders file not found at {processed_orders_path}")
        return
        
    df = pd.read_csv(processed_orders_path)
    
    # 1. Clean and rename fields to match the Raw_Data sheet request
    df_raw = df[[
        "order_id", 
        "customer_id", 
        "order_purchase_timestamp", 
        "customer_state", 
        "promotion_id", 
        "promotion_flag", 
        "order_value", 
        "repeat_customer_flag"
    ]].copy()
    
    df_raw.rename(columns={
        "order_purchase_timestamp": "purchase_date",
        "order_value": "payment_value"
    }, inplace=True)
    
    wb = Workbook()
    
    # --- Style definitions ---
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    soft_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    # --- SHEET 1: Raw_Data ---
    ws_raw = wb.active
    ws_raw.title = "Raw_Data"
    ws_raw.views.sheetView[0].showGridLines = True
    
    for r in dataframe_to_rows(df_raw, index=False, header=True):
        ws_raw.append(r)
        
    # Style Raw_Data headers
    for col in range(1, 9):
        cell = ws_raw.cell(row=1, column=col)
        cell.fill = navy_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")
        
    print("Populated Sheet 1: Raw_Data")
    
    # --- SHEET 2: Pivot_Tables ---
    ws_pivot = wb.create_sheet(title="Pivot_Tables")
    ws_pivot.views.sheetView[0].showGridLines = True
    
    # Helper to write summary sections representing Pivot tables
    def write_pivot_table(ws, title, df_summary, start_row, start_col):
        ws.cell(row=start_row, column=start_col, value=title).font = bold_font
        
        # Headers
        headers = list(df_summary.columns)
        ws.cell(row=start_row+1, column=start_col, value=df_summary.index.name or "Key").font = white_font
        ws.cell(row=start_row+1, column=start_col, value=df_summary.index.name or "Key").fill = navy_fill
        for idx, h in enumerate(headers):
            cell = ws.cell(row=start_row+1, column=start_col+1+idx, value=h)
            cell.font = white_font
            cell.fill = navy_fill
            cell.alignment = Alignment(horizontal="right")
            
        # Data
        curr_row = start_row + 2
        for key, row_vals in df_summary.iterrows():
            ws.cell(row=curr_row, column=start_col, value=key).font = regular_font
            ws.cell(row=curr_row, column=start_col).border = thin_border
            for idx, val in enumerate(row_vals):
                cell = ws.cell(row=curr_row, column=start_col+1+idx, value=val)
                cell.font = regular_font
                cell.border = thin_border
                # Format numbers
                if "AOV" in headers[idx] or "Revenue" in headers[idx]:
                    cell.number_format = "$#,##0.00"
                elif "Rate" in headers[idx] or "%" in headers[idx]:
                    cell.number_format = "0.00%"
                elif "Count" in headers[idx] or "Orders" in headers[idx]:
                    cell.number_format = "#,##0"
            curr_row += 1
            
    # Compute Aggregations
    # 1. Revenue and AOV by Promotion
    promo_agg = df_raw.groupby("promotion_id").agg(
        Revenue=("payment_value", "sum"),
        AOV=("payment_value", "mean"),
        Repeat_Rate=("repeat_customer_flag", "mean")
    )
    
    # 2. Revenue by State
    state_rev = df_raw.groupby("customer_state").agg(
        Revenue=("payment_value", "sum")
    ).sort_values("Revenue", ascending=False).head(10)
    
    # 3. Monthly Revenue Trend
    df_raw["month"] = pd.to_datetime(df_raw["purchase_date"]).dt.strftime("%Y-%m")
    monthly_rev = df_raw.groupby("month").agg(
        Revenue=("payment_value", "sum")
    ).sort_index()
    
    write_pivot_table(ws_pivot, "Revenue, AOV & Repeat Rate by Promotion", promo_agg, 2, 2)
    write_pivot_table(ws_pivot, "Top 10 States by Revenue", state_rev, 2, 7)
    write_pivot_table(ws_pivot, "Monthly Revenue Trend", monthly_rev, 2, 11)
    
    print("Populated Sheet 2: Pivot_Tables")
    
    # --- SHEET 3: Dashboard ---
    ws_dash = wb.create_sheet(title="Dashboard")
    ws_dash.views.sheetView[0].showGridLines = True
    
    # Dashboard Header Block
    ws_dash.cell(row=2, column=2, value="Retail Promotion Analytics Dashboard").font = title_font
    ws_dash.cell(row=3, column=2, value="Campaign Efficacy & Customer Conversion Report").font = Font(name="Calibri", size=11, italic=True)
    
    # Style KPI card helper
    def create_kpi_card(ws, label, value, fmt, start_row, start_col):
        # Merge cells for styling
        for r in range(start_row, start_row+3):
            for c in range(start_col, start_col+3):
                cell = ws.cell(row=r, column=c)
                cell.fill = soft_fill
                cell.border = thin_border
                
        ws.cell(row=start_row, column=start_col, value=label).font = Font(name="Calibri", size=9, bold=True, color="595959")
        ws.cell(row=start_row, column=start_col).alignment = Alignment(horizontal="center", vertical="center")
        
        val_cell = ws.cell(row=start_row+1, column=start_col, value=value)
        val_cell.font = Font(name="Calibri", size=18, bold=True, color="1F497D")
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.number_format = fmt
        
    total_rev = df_raw["payment_value"].sum()
    overall_aov = df_raw["payment_value"].mean()
    overall_ret = df_raw["repeat_customer_flag"].mean()
    promo_rev_pct = df_raw[df_raw["promotion_flag"] == 1]["payment_value"].sum() / total_rev
    
    create_kpi_card(ws_dash, "TOTAL REVENUE", total_rev, "$#,##0", 5, 2)
    create_kpi_card(ws_dash, "AVERAGE ORDER VALUE (AOV)", overall_aov, "$#,##0.00", 5, 6)
    create_kpi_card(ws_dash, "REPEAT PURCHASE RATE", overall_ret, "0.00%", 5, 10)
    create_kpi_card(ws_dash, "PROMO REVENUE SHARE", promo_rev_pct, "0.00%", 5, 14)
    
    # Render mini summary metrics for visualization reference
    ws_dash.cell(row=9, column=2, value="Campaign Comparison Overview").font = bold_font
    ws_dash.cell(row=10, column=2, value="Promotion").font = white_font
    ws_dash.cell(row=10, column=2).fill = navy_fill
    ws_dash.cell(row=10, column=3, value="Revenue").font = white_font
    ws_dash.cell(row=10, column=3).fill = navy_fill
    ws_dash.cell(row=10, column=3).alignment = Alignment(horizontal="right")
    ws_dash.cell(row=10, column=4, value="Repeat Rate").font = white_font
    ws_dash.cell(row=10, column=4).fill = navy_fill
    ws_dash.cell(row=10, column=4).alignment = Alignment(horizontal="right")
    
    curr = 11
    for promo_id, row_v in promo_agg.iterrows():
        ws_dash.cell(row=curr, column=2, value=promo_id).font = regular_font
        ws_dash.cell(row=curr, column=2).border = thin_border
        
        c_rev = ws_dash.cell(row=curr, column=3, value=row_v["Revenue"])
        c_rev.number_format = "$#,##0"
        c_rev.border = thin_border
        
        c_rep = ws_dash.cell(row=curr, column=4, value=row_v["Repeat_Rate"])
        c_rep.number_format = "0.00%"
        c_rep.border = thin_border
        curr += 1
        
    print("Populated Sheet 3: Dashboard Layout Grid")
    
    # Adjust column widths automatically
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            sheet.column_dimensions[col_letter].width = max(max_len + 3, 11)
            
    wb.save(excel_path)
    print(f"Successfully generated clean dashboard workbook at: {excel_path}")

if __name__ == "__main__":
    main()
